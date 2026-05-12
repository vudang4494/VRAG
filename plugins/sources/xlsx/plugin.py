"""Excel/CSV source plugin — uses XlsxChunker for row-aware chunking.

Outputs ParsedDocument with rich metadata (sheet names, column headers, row counts).
The actual format-aware chunking happens in ingestion_v2.py via format_router.
"""
from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import Any, ClassVar

from loguru import logger

from plugins.base import (
    BaseSourcePlugin,
    ParsedDocument,
    PluginCapability,
    PluginConfig,
)


class XlsxSourcePlugin(BaseSourcePlugin):
    """
    Ingest Excel (.xlsx, .xls) and CSV files.
    Returns ParsedDocument with content as markdown-table preview + raw_content for re-chunking.
    """

    name: ClassVar[str] = "xlsx"
    version: ClassVar[str] = "1.0.0"
    capabilities: ClassVar[list[PluginCapability]] = [
        PluginCapability.INGEST_FILE,
        PluginCapability.INGEST_STREAM,
    ]
    supported_types: ClassVar[list[str]] = ["xlsx", "xls", "xlsm", "csv", "tsv"]

    def __init__(self, config: PluginConfig | None = None, credentials: Any = None):
        self.config = config or PluginConfig(raw={})
        self.credentials = credentials

    async def initialize(self) -> None:
        return None

    async def close(self) -> None:
        return None

    async def health_check(self) -> dict[str, Any]:
        return {"status": "ok", "name": self.name, "version": self.version}

    async def fetch(self, url_or_path: str, **kwargs) -> ParsedDocument:
        """
        Fetch via file path or raw bytes (passed in kwargs['content']).

        Returns ParsedDocument:
          - content: markdown-table preview (first sheet, first 50 rows)
          - raw_content: original bytes (for full chunking downstream)
          - metadata: sheet names, total rows, column counts per sheet
        """
        content: bytes | None = kwargs.get("content")
        filename = kwargs.get("filename") or Path(url_or_path).name

        if content is None and url_or_path:
            try:
                with open(url_or_path, "rb") as f:
                    content = f.read()
            except Exception as e:
                logger.error(f"xlsx plugin: cannot read {url_or_path}: {e}")
                raise

        if not content:
            raise ValueError("xlsx plugin: no content provided")

        ext = (Path(filename).suffix or "").lower().lstrip(".")
        is_csv = ext in ("csv", "tsv")

        if is_csv:
            preview, meta = self._preview_csv(content)
        else:
            preview, meta = self._preview_xlsx(content)

        return ParsedDocument(
            title=filename,
            content=preview,
            raw_content=content,
            url=url_or_path,
            file_type="csv" if is_csv else "xlsx",
            file_size_bytes=len(content),
            created_date=datetime.utcnow(),
            metadata={**meta, "filename": filename},
        )

    def _preview_csv(self, content: bytes) -> tuple[str, dict[str, Any]]:
        import csv
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return "(empty CSV)", {"sheet_count": 0, "row_count": 0}
        header, *data = rows
        preview_rows = data[:20]
        md = "| " + " | ".join(header) + " |\n"
        md += "| " + " | ".join(["---"] * len(header)) + " |\n"
        for row in preview_rows:
            md += "| " + " | ".join(c for c in row) + " |\n"
        if len(data) > 20:
            md += f"\n... ({len(data) - 20} more rows)\n"
        return md, {
            "sheet_count": 1,
            "sheets": [{"name": "Sheet1", "row_count": len(data), "column_count": len(header), "columns": header}],
            "row_count": len(data),
        }

    def _preview_xlsx(self, content: bytes) -> tuple[str, dict[str, Any]]:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        except Exception as e:
            logger.warning(f"openpyxl load failed: {e}")
            return f"(unable to read xlsx: {e})", {"sheet_count": 0}

        sheets_meta = []
        preview_parts: list[str] = []
        total_rows = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            it = ws.iter_rows(values_only=True)
            try:
                header_row = next(it)
            except StopIteration:
                continue
            header = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
            data_rows = []
            for row in it:
                if not any(c is not None for c in row):
                    continue
                data_rows.append(list(row))
            sheets_meta.append({
                "name": sheet_name,
                "row_count": len(data_rows),
                "column_count": len(header),
                "columns": header,
            })
            total_rows += len(data_rows)

            # Preview: first 10 rows
            preview_parts.append(f"\n## Sheet: {sheet_name}\n")
            preview_parts.append("| " + " | ".join(header) + " |")
            preview_parts.append("| " + " | ".join(["---"] * len(header)) + " |")
            for r in data_rows[:10]:
                cells = [str(c) if c is not None else "" for c in r]
                preview_parts.append("| " + " | ".join(cells) + " |")
            if len(data_rows) > 10:
                preview_parts.append(f"... ({len(data_rows) - 10} more rows)\n")

        return "\n".join(preview_parts), {
            "sheet_count": len(sheets_meta),
            "sheets": sheets_meta,
            "row_count": total_rows,
        }
