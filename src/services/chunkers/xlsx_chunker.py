"""XLSX / CSV chunker — row-aware, giữ header context cho mỗi chunk.

Strategy:
- 1 sheet → nhiều chunk theo row-group (mỗi chunk = N rows + header).
- Header được prepend trong text mỗi chunk để LLM/embedder hiểu context.
- Metadata: sheet_name, column_names, row_range.
"""

from __future__ import annotations

import csv
import io
from typing import Any

from loguru import logger

from src.services.chunkers.base import BaseChunker, ChunkUnit


class XlsxChunker(BaseChunker):
    name = "xlsx"

    def __init__(
        self,
        rows_per_chunk: int = 20,
        include_header_each_chunk: bool = True,
        **kwargs,
    ):
        super().__init__(**kwargs)
        self.rows_per_chunk = rows_per_chunk
        self.include_header = include_header_each_chunk

    async def chunk(self, content: bytes | str, filename: str = "") -> list[ChunkUnit]:
        is_csv = filename.lower().endswith(".csv")
        if isinstance(content, str):
            content = content.encode("utf-8")

        if is_csv:
            sheets = self._parse_csv(content)
        else:
            sheets = self._parse_xlsx(content)

        if not sheets:
            return []

        units: list[ChunkUnit] = []
        idx = 0
        for sheet_idx, (sheet_name, header, rows) in enumerate(sheets):
            section_parent = None
            if "section" in self.emit_levels and rows:
                section_text = self._format_table(header, rows[:5])
                section_text = (
                    f"[Sheet: {sheet_name}]\n{section_text}\n... ({len(rows)} rows total)"
                )
                units.append(
                    ChunkUnit(
                        text=section_text,
                        chunk_index=idx,
                        chunk_level="section",
                        parent_index=None,
                        metadata={
                            "sheet_name": sheet_name,
                            "column_names": header,
                            "row_count": len(rows),
                            "filename": filename,
                            "format": "csv" if is_csv else "xlsx",
                        },
                    )
                )
                section_parent = idx
                idx += 1

            for i in range(0, len(rows), self.rows_per_chunk):
                row_batch = rows[i : i + self.rows_per_chunk]
                row_range = (i, i + len(row_batch))
                text = self._format_table(header if self.include_header else None, row_batch)
                text = f"[Sheet: {sheet_name}, rows {row_range[0] + 1}-{row_range[1]}]\n{text}"
                units.append(
                    ChunkUnit(
                        text=text,
                        chunk_index=idx,
                        chunk_level="paragraph",
                        parent_index=section_parent,
                        metadata={
                            "sheet_name": sheet_name,
                            "column_names": header,
                            "row_range": list(row_range),
                            "filename": filename,
                            "format": "csv" if is_csv else "xlsx",
                        },
                    )
                )
                idx += 1
        return units

    @staticmethod
    def _format_table(header: list[str] | None, rows: list[list[Any]]) -> str:
        lines = []
        if header:
            lines.append("| " + " | ".join(str(h) for h in header) + " |")
            lines.append("| " + " | ".join(["---"] * len(header)) + " |")
        for row in rows:
            cells = [str(c) if c is not None else "" for c in row]
            lines.append("| " + " | ".join(cells) + " |")
        return "\n".join(lines)

    def _parse_csv(self, content: bytes) -> list[tuple[str, list[str], list[list[Any]]]]:
        try:
            text = content.decode("utf-8", errors="replace")
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if not rows:
                return []
            header, *data = rows
            return [("Sheet1", [h.strip() for h in header], data)]
        except Exception as e:
            logger.warning(f"CSV parse failed: {e}")
            return []

    def _parse_xlsx(self, content: bytes) -> list[tuple[str, list[str], list[list[Any]]]]:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        except Exception as e:
            logger.warning(f"openpyxl parse failed: {e}")
            return []

        result: list[tuple[str, list[str], list[list[Any]]]] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            it = ws.iter_rows(values_only=True)
            try:
                header_row = next(it)
            except StopIteration:
                continue
            header = [
                str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_row)
            ]
            data_rows = [list(row) for row in it if any(c is not None for c in row)]
            if data_rows:
                result.append((sheet_name, header, data_rows))
        return result
